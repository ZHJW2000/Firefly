"""阶段6：整合五阶段结果，风险评分排序，输出 Excel 报表。"""

import os
from urllib.parse import urlparse

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from .risk import score_service


def run(ctx, log, progress, should_stop):
    data = ctx.data
    ports_data = data.get("ports") or []
    probes = {p["url"]: p for p in (data.get("probes") or [])}
    fps = {f["url"]: f for f in (data.get("fingerprints") or [])}
    endpoints = data.get("endpoints") or []
    sensitive = data.get("sensitive") or []

    # 预解析端点，便于按 host+port 归属
    def _hostport(u):
        pr = urlparse(u)
        port = pr.port or (443 if pr.scheme == "https" else 80)
        return pr.hostname, port

    ep_parsed = [(_hostport(u), u) for u in endpoints]
    sens_parsed = [(_hostport(s["url"]), s) for s in sensitive]
    probe_parsed = []
    for u, pr in probes.items():
        h, pt = _hostport(u)
        probe_parsed.append(((h, pt), u, pr, fps.get(u)))

    # 逐服务评分并排序
    rows = []
    for entry in ports_data:
        ip = entry["ip"]
        hosts = entry.get("hosts") or []
        host_set = {ip} | set(hosts)
        for p in entry["ports"]:
            port = p["port"]
            # 探测/指纹记录：IP 或任一域名 + 端口匹配（重定向到其他域名时按端口回退）
            probe, fps_, fp_text = None, None, ""
            for (h, pt), u, pr, fp in probe_parsed:
                if pt == port and (h in host_set or h is None):
                    probe, fps_ = pr, fp
                    if fp:
                        fp_text = " / ".join(x for x in (fp.get("cms"), fp.get("server")) if x)
                    break
            if probe is None:
                for (h, pt), u, pr, fp in probe_parsed:
                    if pt == port:
                        probe, fps_ = pr, fp
                        if fp:
                            fp_text = " / ".join(x for x in (fp.get("cms"), fp.get("server")) if x)
                        break
            n_ep = sum(1 for (h, pt), _ in ep_parsed if pt == port and h in host_set)
            n_sens = sum(1 for (h, pt), s in sens_parsed if pt == port and h in host_set)
            is_web = bool(probe) or port in (80, 443, 8080, 8443, 8000, 8008, 8888, 7001, 9090)
            score, reasons = score_service(port, p.get("service", ""), fp_text, n_sens, is_web)
            probe = probe or {}
            rows.append({
                "score": score, "reasons": "; ".join(reasons),
                "ip": ip, "host": hosts[0] if hosts else "", "port": port,
                "service": p.get("service", ""), "finger": fp_text,
                "status": probe.get("status", ""), "title": probe.get("title", "") or (fps_ or {}).get("title", ""),
                "url": probe.get("url", "") if probe else "",
                "n_ep": n_ep, "n_sens": n_sens,
            })
    rows.sort(key=lambda r: (-r["score"], r["ip"], r["port"]))
    log(f"共 {len(rows)} 个服务，最高风险 {rows[0]['score'] if rows else 0} 分。")

    targets = ctx.targets or [ctx.target]
    if len(targets) == 1:
        report = os.path.join(ctx.outdir, f"资产测绘报告_{ctx.target}.xlsx")
    else:
        report = os.path.join(ctx.outdir, f"资产测绘报告_批量{len(targets)}目标.xlsx")
    _write_excel(report, targets, rows, endpoints, sensitive, ports_data,
                 data.get("subdomains") or [], probes)
    log("Excel 报表生成完成。")
    return {"data": {"report_path": report, "services": rows}}


def _write_excel(path, targets, rows, endpoints, sensitive, ports_data, subdomains, probes):
    wb = Workbook()
    hdr_font = Font(bold=True, color="FFFFFF")
    hdr_fill = PatternFill("solid", fgColor="4472C4")
    risk_fill = [(60, PatternFill("solid", fgColor="F8CBAD")),
                 (30, PatternFill("solid", fgColor="FFE699")),
                 (0, PatternFill("solid", fgColor="C6E0B4"))]

    def style(ws, widths):
        for col, w in enumerate(widths, 1):
            c = ws.cell(row=1, column=col)
            c.font, c.fill = hdr_font, hdr_fill
            ws.column_dimensions[get_column_letter(col)].width = w
        ws.freeze_panes = "A2"

    # Sheet1 高危服务总览（评分降序）
    ws = wb.active
    ws.title = "高危服务总览"
    cols = ["风险分", "风险原因", "IP", "域名", "端口", "服务", "指纹/组件", "状态码", "Title", "URL", "API端点数", "敏感数据数"]
    ws.append(cols)
    for r in rows:
        ws.append([r["score"], r["reasons"], r["ip"], r["host"], r["port"], r["service"],
                   r["finger"], r["status"], r["title"], r["url"], r["n_ep"], r["n_sens"]])
    style(ws, [8, 30, 14, 22, 7, 10, 24, 8, 28, 30, 10, 10])
    for row in range(2, ws.max_row + 1):
        v = ws.cell(row=row, column=1).value or 0
        for threshold, fill in risk_fill:
            if v >= threshold:
                ws.cell(row=row, column=1).fill = fill
                break

    # Sheet2 敏感信息明细
    ws = wb.create_sheet("敏感信息明细")
    ws.append(["来源URL", "规则", "证据片段"])
    for s in sensitive:
        ws.append([s["url"], s["rule"], s["evidence"]])
    style(ws, [50, 22, 70])

    # Sheet3 API 端点清单
    ws = wb.create_sheet("API端点清单")
    ws.append(["端点URL"])
    for u in endpoints:
        ws.append([u])
    style(ws, [90])

    # Sheet4 端口/服务清单
    ws = wb.create_sheet("端口服务清单")
    ws.append(["IP", "域名", "端口", "服务"])
    for entry in ports_data:
        for p in entry["ports"]:
            ws.append([entry["ip"], ", ".join(entry.get("hosts") or []), p["port"], p["service"]])
    style(ws, [15, 30, 8, 14])

    # Sheet5 子域清单
    ws = wb.create_sheet("子域清单")
    ws.append(["子域", "解析IP"])
    for s in subdomains:
        ws.append([s])
    style(ws, [40, 16])

    # 顶部说明页
    ws = wb.create_sheet("评估说明", 0)
    ws["A1"] = "资产测绘报告"
    ws["A1"].font = Font(bold=True, size=14)
    info = [
        ("目标", "、".join(targets) if len(targets) <= 10
         else f"{targets[0]} 等 {len(targets)} 个目标"),
        ("目标数量", len(targets)),
        ("子域数", len(subdomains)),
        ("扫描IP数", len(ports_data)),
        ("Web存活URL数", len(probes)),
        ("API端点数", len(endpoints)),
        ("敏感数据命中", len(sensitive)),
        ("服务总数", len(rows)),
        ("高危服务(≥60分)", sum(1 for r in rows if r["score"] >= 60)),
        ("说明", "本报告由授权范围内的资产测绘流程生成，仅含公开可访问信息，不含任何凭据内容。"),
    ]
    for i, (k, v) in enumerate(info, start=3):
        ws.cell(row=i, column=1, value=k).font = Font(bold=True)
        ws.cell(row=i, column=2, value=v)
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 70
    wb.save(path)
