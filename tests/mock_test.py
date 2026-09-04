"""AssetMapper mock 自测：不依赖外部安全工具，验证解析器/探测/评分/报表。

运行: python tests/mock_test.py
"""

import http.server
import json
import os
import sys
import tempfile
import threading

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from assetmap import risk, s1_subdomain, s2_portscan, s5_crawl
from assetmap.pipeline import Context
from assetmap.s4_probe import probe_urls
from assetmap.s6_report import run as s6_run

INDEX = "<html><head><title>Mock 管理后台</title></head><body><h1>ok</h1></body></html>"
APP_JS = """var cfg = {password: 'SuperSecret@2024', access_key: 'LTAI5tAbCdEfGh12345'};
var api = '/api/v1/users'; var db = 'jdbc:mysql://10.0.0.5:3306/school';"""


class Handler(http.server.BaseHTTPRequestHandler):
    PAGES = {"/": INDEX, "/static/app.js": APP_JS}

    def do_GET(self):
        body = self.PAGES.get(self.path)
        if body is None:
            self.send_response(404); self.end_headers(); return
        data = body.encode()
        self.send_response(200)
        self.send_header("Content-Type", "text/html; charset=utf-8")
        self.send_header("Server", "nginx/1.18.0")
        self.send_header("Content-Length", str(len(data)))
        self.end_headers()
        self.wfile.write(data)

    def log_message(self, *a):
        pass


def start_server():
    srv = http.server.ThreadingHTTPServer(("127.0.0.1", 0), Handler)
    threading.Thread(target=srv.serve_forever, daemon=True).start()
    return srv, srv.server_address[1]


def test_s1_validate():
    subs = s1_subdomain._validate(
        ["a.example.com", "a.example.com", "b.example.com.", "BAD!!", "evil.com", "aexample.com"],
        "example.com", lambda m: None)
    assert subs == ["a.example.com", "b.example.com"], subs
    # 中文 edu.cn 三级域场景
    subs2 = s1_subdomain._validate(["www.xjtu.edu.cn", "lib.xjtu.edu.cn", "x.com"],
                                   "xjtu.edu.cn", lambda m: None)
    assert subs2 == ["www.xjtu.edu.cn", "lib.xjtu.edu.cn"], subs2
    print("s1 域名校验通过")


def test_s2_parse():
    gnmap = """# Nmap 7.80 scan initiated
Host: 1.2.3.4 (a.example.com)\tStatus: Up
Host: 1.2.3.4 (a.example.com)\tPorts: 22/open/tcp//ssh//OpenSSH 7.4/, 80/open/tcp//http//nginx/, 3306/open/tcp//mysql//MySQL 5.7/
Host: 5.6.7.8 ()\tPorts: 443/open/tcp//https//, 8443/open/tcp//https-alt//
"""
    tmp = tempfile.mktemp(suffix=".gnmap")
    open(tmp, "w").write(gnmap)
    result = s2_portscan._parse_gnmap(tmp, {"a.example.com": "1.2.3.4", "b.example.com": "5.6.7.8"})
    os.remove(tmp)
    assert len(result) == 2
    r1 = next(r for r in result if r["ip"] == "1.2.3.4")
    assert [p["port"] for p in r1["ports"]] == [22, 80, 3306]
    assert r1["ports"][0]["service"] == "ssh" and r1["hosts"] == ["a.example.com"]
    print("s2 gnmap 解析通过")


def test_s4_probe(srv):
    port = srv.server_address[1]
    try:
        rs = probe_urls([f"http://127.0.0.1:{port}/", f"http://127.0.0.1:{port}/404",
                         f"http://127.0.0.1:1/"], concurrency=10, timeout=5)
        assert len(rs) == 2, rs  # 200 与 404 都算存活；死端口不返回
        r = next(x for x in rs if x["status"] == 200)
        assert r["status"] == 200 and r["title"] == "Mock 管理后台" and r["headers"].get("Server")
        print("s4 探测通过:", r["url"], r["status"], r["title"])
        return f"http://127.0.0.1:{port}"
    except AssertionError:
        raise


def test_s5_scan(base):
    # JS 扫描对 mock 内容命中：硬编码密码、云密钥、数据库连接串、内网IP
    findings = s5_crawl._scan_js([f"{base}/static/app.js"], lambda m: None,
                                 lambda d, t: None, lambda: False)
    rules = {f["rule"] for f in findings}
    assert "硬编码密码" in rules, rules
    assert "硬编码云密钥(AK/SK)" in rules or "API Key 通用格式" in rules, rules
    assert "数据库连接串" in rules, rules
    assert "内网 IP 暴露" in rules, rules
    print("s5 敏感数据扫描通过:", rules)


def test_s6_report():
    ports_data = [{"ip": "1.2.3.4", "hosts": ["a.example.com"],
                   "ports": [{"port": 6379, "service": "redis"}, {"port": 8080, "service": "http"}]}]
    probes = [{"url": "http://1.2.3.4:8080", "status": 200, "title": "Spring Boot 管理平台",
               "length": 100, "headers": {"Server": "Tomcat"}}]
    fps = [{"url": "http://1.2.3.4:8080", "cms": "Spring Boot", "server": "Tomcat",
            "status": 200, "title": "Spring Boot 管理平台"}]
    sensitive = [{"url": "http://1.2.3.4:8080/static/app.js", "rule": "硬编码密码", "evidence": "xxx"}]
    endpoints = ["http://1.2.3.4:8080/api/users", "http://1.2.3.4:8080/static/app.js"]

    outdir = tempfile.mkdtemp()
    ctx = Context(target="example.com", outdir=outdir, cfg={})
    ctx.data.update({"dns": {"a.example.com": "1.2.3.4"},
                     "ports": ports_data, "probes": probes, "fingerprints": fps,
                     "sensitive": sensitive, "endpoints": endpoints,
                     "subdomains": ["a.example.com"]})
    payload = s6_run(ctx, lambda m: None, lambda d, t: None, lambda: False)
    rows = payload["data"]["services"]
    assert len(rows) == 2
    top = rows[0]
    assert top["port"] == 8080 and top["score"] >= 40, top  # spring boot(15)+tomcat(10)+敏感(12)+非标(5) = 42
    report = payload["data"]["report_path"]
    assert os.path.isfile(report) and os.path.getsize(report) > 5000

    from openpyxl import load_workbook
    wb = load_workbook(report)
    assert wb.sheetnames[0] == "评估说明" and "高危服务总览" in wb.sheetnames, wb.sheetnames
    summary = {r[0].value: r[1].value for r in wb["评估说明"].iter_rows(min_row=3, max_row=11)}
    assert summary["服务总数"] == 2, summary
    print("s6 报表通过，8080 评分 =", top["score"], "原因:", top["reasons"])


def test_risk():
    s, r = risk.score_service(6379, "redis", "", 0, False)
    assert s >= 20 and "Redis" in " ".join(r)
    s2, _ = risk.score_service(8080, "http", "Shiro", 2, True)
    assert s2 >= 25 + 24 + 5 - 5, s2  # shiro + 2条敏感(封顶30取24? 每条12->24) + 非标
    print("risk 评分通过")


if __name__ == "__main__":
    test_s1_validate()
    test_s2_parse()
    test_risk()
    srv, _ = start_server()          # 服务器贯穿 s4/s5 测试
    base = test_s4_probe(srv)
    test_s5_scan(base)
    srv.shutdown()
    test_s6_report()
    print("\n=== AssetMapper 全部 mock 自测通过 ===")
