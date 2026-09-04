"""端到端：用 Pipeline 状态机对 127.0.0.1 跑完整六阶段（阶段1 用手工子域列表）。"""

import http.server
import os
import shutil
import sys
import threading

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
os.chdir(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from assetmap.pipeline import Context, Pipeline, STAGES

INDEX = "<html><head><title>E2E Mock Portal</title></head><body><a href='/static/app.js'>app</a></body></html>"
APP_JS = "var cfg={password:'E2E@Pass123',db:'jdbc:mysql://10.0.0.9:3306/core'};"


class Handler(http.server.BaseHTTPRequestHandler):
    PAGES = {"/": INDEX, "/static/app.js": APP_JS}

    def do_GET(self):
        body = self.PAGES.get(self.path)
        if body is None:
            self.send_response(404); self.end_headers(); return
        data = body.encode()
        self.send_response(200)
        self.send_header("Content-Type", "text/html; charset=utf-8")
        self.send_header("Content-Length", str(len(data)))
        self.end_headers()
        self.wfile.write(data)

    def log_message(self, *a):
        pass


def main():
    import socket
    port = None
    for p in (8000, 8080, 8888):
        if p == 8000:
            pass
        try:
            s = socket.socket(); s.bind(("127.0.0.1", p)); s.close(); port = p; break
        except OSError:
            continue
    srv = http.server.ThreadingHTTPServer(("127.0.0.1", port), Handler)
    threading.Thread(target=srv.serve_forever, daemon=True).start()
    print(f"mock: http://127.0.0.1:{port}")

    outdir = os.path.abspath("output/_e2e")
    shutil.rmtree(outdir, ignore_errors=True)
    cfg = {
        "oneforall_python": "",  # 用手工子域列表
        "oneforall_py": "",
        "nmap_exe": r"E:\Cybersecurity\tools\01-Information-Gathering\Nmap\nmap.exe",
        "nmap_mode": "top1000",
        "ehole_exe": r"E:\Cybersecurity\One-Fox\tools\gui_scan\ehole\EHole_windows_amd64.exe",
        "ehole_cwd": r"E:\Cybersecurity\One-Fox\tools\gui_scan\ehole",
        "katana_exe": r"E:\Cybersecurity\tools\01-Information-Gathering\katana\katana.exe",
        "probe_concurrency": 50, "katana_depth": 1,
    }
    ctx = Context(target="e2e.local", outdir=outdir, cfg=cfg)
    # 阶段1降级：手工子域列表（等价 GUI 导入）
    ctx.data["manual_subdomains"] = ["localhost"]
    statuses = {}
    pipe = Pipeline(ctx, log=lambda m: print("   ", m),
                    progress=lambda i, d, t: None,
                    stage_status=lambda i, s: statuses.__setitem__(i, s))
    data = pipe.run()
    print("阶段状态:", {STAGES[i-1][0]: s for i, s in statuses.items()})
    assert data.get("report_path"), "未生成报表"
    assert os.path.isfile(data["report_path"])

    from openpyxl import load_workbook
    wb = load_workbook(data["report_path"])
    print("工作表:", wb.sheetnames)
    rows = list(wb["高危服务总览"].iter_rows(min_row=2, values_only=True))
    mock = next((r for r in rows if r[4] == port), None)
    assert mock, f"报表未包含 mock 端口 {port}: {rows[:3]}"
    print(f"mock 服务: 端口={mock[4]} 风险分={mock[0]} title={mock[8]} 端点数={mock[10]} 敏感数={mock[11]}")
    assert mock[11] >= 1, "应有敏感数据命中"
    srv.shutdown()
    print("\n=== 端到端测试通过 ===")


if __name__ == "__main__":
    main()
